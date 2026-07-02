"""One-off: inspect QazaqTest_catalog.xlsx. Run: python scripts/inspect_catalog_xlsx.py [path]"""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl


def main() -> None:
    p = Path(sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\user\Downloads\QazaqTest_catalog.xlsx")
    if not p.exists():
        print(json.dumps({"error": "file_not_found", "path": str(p)}, ensure_ascii=False))
        sys.exit(1)

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    out: dict = {"path": str(p), "sheets": wb.sheetnames}

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[name] = {"empty": True}
            continue
        hdr = list(rows[0])
        out[name] = {
            "nrows": len(rows),
            "ncols": max(len(r) for r in rows),
            "header": hdr,
            "sample_row": list(rows[1]) if len(rows) > 1 else None,
        }

    if "Каталог" in wb.sheetnames:
        ws = wb["Каталог"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = list(rows[0])
        idx = {h: i for i, h in enumerate(hdr) if h is not None}

        crm_cats = Counter()
        orig_cats = Counter()
        subcats = Counter()
        site_ids = Counter()

        for r in rows[1:]:
            if not r or all(x is None for x in r):
                continue
            for key in ("Категория (CRM)", "Категория", "Категория сайта"):
                if key in idx and r[idx[key]] is not None:
                    crm_cats[str(r[idx[key]]).strip()] += 1
                    break
            if "Оригинальная категория" in idx and r[idx["Оригинальная категория"]] is not None:
                orig_cats[str(r[idx["Оригинальная категория"]]).strip()] += 1
            if "Подкатегория" in idx and r[idx["Подкатегория"]] is not None:
                subcats[str(r[idx["Подкатегория"]])] += 1
            for key in ("categoryId", "ID категории сайта", "Сайт categoryId"):
                if key in idx and r[idx[key]] is not None:
                    site_ids[str(r[idx[key]]).strip()] += 1
                    break

        out["_catalog_stats"] = {
            "data_rows": len([x for x in rows[1:] if x and any(c is not None for c in x)]),
            "crm_categories": crm_cats.most_common(40),
            "original_categories_top": orig_cats.most_common(25),
            "subcategories_top": subcats.most_common(15) if subcats else None,
            "site_category_ids": site_ids.most_common(20) if site_ids else None,
        }

    wb.close()
    text = json.dumps(out, ensure_ascii=False, indent=2)
    out_path = Path(__file__).resolve().parent / "_catalog_inspect_out.json"
    out_path.write_text(text, encoding="utf-8")
    print(out_path)


if __name__ == "__main__":
    main()
